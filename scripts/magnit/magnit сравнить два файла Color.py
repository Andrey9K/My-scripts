from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook
from openpyxl.utils import units


wb = load_workbook(f'/home/akozlov/Документы/Project/Marvel/pimcore/pimcore/public/excelTemplates/magnit1p/magnit1p.xlsx')
wb2 = load_workbook(f'/home/akozlov/Загрузки/111/magnit1p.xlsx')
sheet = wb2['Цвета']
sheet2 = wb['Цвета']
# mylist1=[]
dates = [x.value for x in sheet.iter_cols(min_row=1,min_col=1,max_col=1).send(None)]
print(dates)
# for hed in dates:
#     mylist1.append(hed) 
    
# mylist2=[]
dates2 = [x.value for x in sheet2.iter_cols(min_row=1,min_col=1,max_col=1).send(None)]
print(dates2)
# for hed in dates2:
#     mylist2.append(hed) 
# print(i)

result = list(set(dates) - set(dates2))
result1 = list(set(dates2) - set(dates))
# if 'Артикул производителя (Part Number)*' in result:
#     print(i)
print('Отличие Старого от Нового:')
for i in result:    
    print(f'- {i}')

print('Отличие Нового от Старого:')
for p in result1:    
    print(f'- {p}')
# print(f'Отличие Старого от Нового: {result}')
# print(f'Отличие Нового от Старого: {result1}')