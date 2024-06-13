from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(f'/home/akozlov/Документы/Project/Marvel/pimcore/pimcore/public/excelTemplates/magnit1p/magnit1p.xlsx')
wb2 = load_workbook(f'/home/akozlov/Загрузки/111/magnit1pold.xlsx')
sheet = wb2['Категории']
sheet2 = wb['Категории']
dictionary = {}
dictionary2 = {}
notresult=[]

uselist=[14504, 13827, 14281, 11844, 13997, 13955, 14282, 13898, 11770, 13507, 12154, 12556, 11718, 13716, 13401, 14373, 12585, 12041, 13329, 12608, 14726, 12194, 13610, 12465, 11558, 13060, 11772, 13601, 13448, 12383, 13086, 14035, 14232, 12894, 12684, 12165, 11595, 12185, 12032, 11906, 2573, 2572, 14081, 13186, 11801, 11859, 12877, 13265, 12547, 12587, 2719, 14594, 14595, 14596, 13815, 12756, 14085, 12668, 12850, 11951, 13862, 14766, 12377, 11754, 14537, 12020, 14768, 14767, 14801, 14770, 14769, 12416, 14771, 12173, 13810, 14778, 12691, 13025, 12232, 13460, 13516, 13194, 14214, 13067, 11685, 12137, 11905, 14015, 12821, 12504, 12538, 13837, 13992, 12261, 12951, 13677, 12049, 12964, 12222, 12952, 14037, 12243, 14065, 12381, 13844, 12157, 12327, 12270, 13236, 13715, 13076, 14140, 13187, 13830, 12092, 12977, 11645, 13520, 13259, 12324, 13238, 14561, 13912, 13942, 13221, 12152, 11605, 13222, 13786, 12941, 13658, 13083, 13425, 12980, 12721, 2803, 14277, 13100, 2800, 13948, 12045, 12520, 12696, 12418, 13548, 11543, 12184, 12252, 13364, 11765, 13871, 13978, 11620, 11820, 13510, 13779, 12781, 12171, 14225, 12615, 11937, 13626, 14262, 12690, 12601, 14502, 2802, 13381, 11939, 12542, 12544, 13945, 14266, 12339, 14199, 14110, 12635, 12843, 14136, 12665, 2892, 11631, 13880]
    
# Получить максимальное количество заполненных ячеек в первом столбце
max_row = sheet.max_row
max_row2 = sheet2.max_row
# Пройти по каждой строке и добавить значения в словарь
for row in range(2, max_row+1):
        if sheet[f'B{row}'].value != None:
            cell_b = sheet[f'B{row}'].value             
            cell_d = sheet[f'C{row}'].value   
        # Добавить значения в словарь
        dictionary[cell_b] =cell_d
for row2 in range(2, max_row2+1):
        if sheet2[f'B{row2}'].value != None:
            cell_b = sheet2[f'B{row2}'].value            
            cell_d = sheet2[f'C{row2}'].value   
        # Добавить значения в словарь
        dictionary2[cell_b] = cell_d

def find_differences(dictionary, dictionary2):
# Нахождение ключей, которые присутствуют только в одном из словарей
    keys_only_in_dict1=[]
    keys_only_in_dict2=[]
    keys_only_in_dict1 = list(set(dictionary.keys()) - set(dictionary2.keys()))
    keys_only_in_dict2 = list(set(dictionary2.keys()) - set(dictionary.keys()))

# Нахождение различий по значениям для общих ключей
    value_differences = {}
    common_keys = set(dictionary.keys()).intersection(dictionary2.keys())
    for key in common_keys:
        value1 = dictionary[key]
        value2 = dictionary2[key]
        if value1 != value2:
            value_differences[key] = (value1, value2)

    return keys_only_in_dict1, keys_only_in_dict2, value_differences

# Получение различий между словарями
keys_only_in_dict1, keys_only_in_dict2, value_differences = find_differences(dictionary, dictionary2)

# Вывод результатов
if keys_only_in_dict2 != [] or keys_only_in_dict1 != [] or value_differences !={}:
        
    # if keys_only_in_dict2 != []:
    #     notresult.append('NOK')
    #     print('Отсутствуют:\n')   
       
    #     for key, values in dictionary2.items():
    #         #  print(key)
    #          if key in keys_only_in_dict2:
    #               print(values,key)  
    #     for u in keys_only_in_dict2:
    #         print(u)
    
    # if keys_only_in_dict1 != []:
    #     notresult.append('NOK')
    #     print('\nЛишние:\n')
    #     print(keys_only_in_dict1)
    #     for u1 in keys_only_in_dict1:        
    #         print(u1)
    
    if value_differences !={}:
        notresult.append('NOK')
        print("Различия по category_id, Название категории:")
        for key, values in value_differences.items():
            if key in uselist:
                value1, value2 = values
                print(f"id атрибута: {key}, Значение файле1: {value1}, Значение файле2: {value2}")
    else: 
        print('Различий по параметрам нет')

if notresult == []:
      print('Различий не обнаруженно')      

