from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(f'/home/akozlov/Документы/Project/Marvel/pimcore/pimcore/public/excelTemplates/magnit1p/magnit1p.xlsx')
wb2 = load_workbook(f'/home/akozlov/Загрузки/111/magnit1p.xlsx')
sheet = wb2['Категории']
sheet2 = wb['Категории']
dictionary = {}
dictionary2 = {}
notresult=[]

    
# Получить максимальное количество заполненных ячеек в первом столбце
max_row = sheet.max_row
max_row2 = sheet2.max_row
# Пройти по каждой строке и добавить значения в словарь
for row in range(2, max_row+1):
        if sheet[f'B{row}'].value != None:
            cell_b = sheet[f'B{row}'].value 
            cell_c = sheet[f'A{row}'].value
            cell_d = sheet[f'C{row}'].value   
        # Добавить значения в словарь
        dictionary[cell_b] = [cell_c, cell_d]
for row2 in range(2, max_row2+1):
        if sheet2[f'B{row2}'].value != None:
            cell_b = sheet2[f'B{row2}'].value 
            cell_c = sheet2[f'A{row2}'].value
            cell_d = sheet2[f'C{row2}'].value   
        # Добавить значения в словарь
        dictionary2[cell_b] = [cell_c, cell_d]

def find_differences(dictionary, dictionary2):
# Нахождение ключей, которые присутствуют только в одном из словарей
    keys_only_in_dict1=[]
    keys_only_in_dict2=[]
    keys_only_in_dict1 = list(set(dictionary.keys()) - set(dictionary2.keys()))
    keys_only_in_dict2 = list(set(dictionary2.keys()) - set(dictionary.keys()))

# Нахождение различий по значениям для общих ключей
    value_differences = {}
    common_keys = set(dictionary.keys()).intersection(dictionary2.keys())
    for key in common_keys:
        value1 = dictionary[key]
        value2 = dictionary2[key]
        if value1 != value2:
            value_differences[key] = (value1, value2)

    return keys_only_in_dict1, keys_only_in_dict2, value_differences

# Получение различий между словарями
keys_only_in_dict1, keys_only_in_dict2, value_differences = find_differences(dictionary, dictionary2)

# Вывод результатов
if keys_only_in_dict2 != [] or keys_only_in_dict1 != [] or value_differences !={}:
        
    if keys_only_in_dict2 != []:
        notresult.append('NOK')
        print('Отсутствуют:\n')   
       
        for key, values in dictionary2.items():
            #  print(key)
             if key in keys_only_in_dict2:
                  print(values,key)  
        # for u in keys_only_in_dict2:
        #     print(u)
    
    # if keys_only_in_dict1 != []:
    #     notresult.append('NOK')
    #     print('\nЛишние:\n')
    #     print(keys_only_in_dict1)
    #     for u1 in keys_only_in_dict1:        
    #         print(u1)
    
    # if value_differences !={}:
    #     notresult.append('NOK')
    #     print("Различия по category_id, Название категории, category.string_path:")
    #     for key, values in value_differences.items():
    #         value1, value2 = values
    #         print(f"id атрибута: {key}, Значение файле1: {value1}, Значение файле2: {value2}")
    # else: 
    #     print('Различий по параметрам нет')

if notresult == []:
      print('Различий не обнаруженно')      

