from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import json
# Установка дополнительный библиотек
# pip install openpyxl
# pip install requests

# Укажите путь к файлу для анализа
pathfile= '/home/akozlov/Загрузки/Analitika OZON.xlsx'

# Укажите количество листов для анализа
numlist=120

# Открывается файл и запускается цикл
wb = load_workbook(pathfile)
notresult=[]
for i in wb.sheetnames[0:numlist]:       
    dictionary = {}
    sheet = wb[f'{i}']
    
    # Получить максимальное количество заполненных ячеек в первом столбце
    max_row = sheet.max_row

    # Пройти по каждой строке и добавить значения в словарь
    for row in range(5, max_row+1):
            if sheet[f'B{row}'].value != None:
                cell_b = sheet[f'B{row}'].value 
                cell_c = sheet[f'C{row}'].value
                cell_d = sheet[f'D{row}'].value
                cell_e = sheet[f'E{row}'].value
                cell_f = sheet[f'F{row}'].value
                cell_g = sheet[f'G{row}'].value

            # Добавить значения в словарь
            dictionary[cell_b] = [cell_c, cell_d, cell_e, cell_f, cell_g]
    url = "https://api-seller.ozon.ru/v1/description-category/attribute"
    category_id= sheet['F2'].value
    type_id=sheet['F3'].value
    


    payload = json.dumps({
    "description_category_id": f'{category_id}',
    "type_id": f'{type_id}',
    "language": "DEFAULT"
    })
    headers = {
    'Client-Id': '77899',
    'Api-Key': '67de6a35-eaaf-4b08-ba84-62f525a6e0c8',
    'Content-Type': 'text/plain',
    'Cookie': '__cf_bm=zBIfnu0fjArVn6K_OsxSXfq.zy8lDVG5tcuIFOkS02w-1698179182-0-AanjbwikPZTpzfRXbf5x4GHbDIuhOSxpqd4Ewr80udCehPwPgvlmMiIREm6xvkoan2HVPZV7J0oOg+exMi1or9U='
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    data=response.json()
    templates = data
    
    js1={}
    if   'result' in templates: 
        if isinstance(templates['result'], list):
                    s= templates['result']
                    for t in range(0,len(s)):            
                        s1= s[t]
                        s2=s1['id']
                        s3=s1['name']
                        s4=s1['type']
                        s5=s1['is_required']
                        if s5 == False:
                              s5='ЛОЖЬ'
                        if s5 == True:
                              s5='ИСТИНА'
                        s6=s1['is_collection']
                        if s6 == False:
                              s6='ЛОЖЬ'
                        if s6 == True:
                              s6='ИСТИНА'
                        s7=s1['dictionary_id']
                        js1[s2]=[s3, s4, s5, s6, s7]
        else: 
            print(f'{i} : Не корректный Json')
    else:
          print(f'{i} : Не корректный Json')
   
    def find_differences(dictionary, js1):
    # Нахождение ключей, которые присутствуют только в одном из словарей
        keys_only_in_dict1=[]
        keys_only_in_dict2=[]
        keys_only_in_dict1 = list(set(dictionary.keys()) - set(js1.keys()))
        keys_only_in_dict2 = list(set(js1.keys()) - set(dictionary.keys()))

    # Нахождение различий по значениям для общих ключей
        value_differences = {}
        common_keys = set(dictionary.keys()).intersection(js1.keys())
        for key in common_keys:
            value1 = dictionary[key]
            value2 = js1[key]
            if value1 != value2:
                value_differences[key] = (value1, value2)

        return keys_only_in_dict1, keys_only_in_dict2, value_differences

    # Получение различий между словарями
    keys_only_in_dict1, keys_only_in_dict2, value_differences = find_differences(dictionary, js1)

    # Вывод результатов
    if keys_only_in_dict2 != [] or keys_only_in_dict1 != [] or value_differences !={}:
        print(i)
        
        if keys_only_in_dict2 != []:
            notresult.append('NOK')
            print('Отсутствуют:\n')            
            for u in keys_only_in_dict2:
                print(u)
        
        if keys_only_in_dict1 != []:
            notresult.append('NOK')
            print('\nЛишние:\n')
            print(keys_only_in_dict1)
            for u1 in keys_only_in_dict1:        
                print(u1)
        
        if value_differences !={}:
            notresult.append('NOK')
            print("Различия по name, type, is_required, is_collection, dictionary_id:")
            for key, values in value_differences.items():
                value1, value2 = values
                print(f"id атрибута: {key}, Значение файле: {value1}, Значение в OZON: {value2}")
        else: 
            print('Различий по параметрам нет')

if notresult == []:
      print('Изменений в API OZON НЕТ')      

