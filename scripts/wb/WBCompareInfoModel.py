from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import json
# Установка дополнительный библиотек
# pip install openpyxl
# pip install requests

# Укажите путь к файлу для анализа
pathfile= '/home/akozlov/Загрузки/Аналитика WB.xlsx'

# Укажите количество листов для анализа
numlist=126

# Открывается файл и запускается цикл
wb = load_workbook(pathfile)
notresult=[]
for i in wb.sheetnames[1:numlist]:       
    dictionary = {}
    sheet = wb[f'{i}']
    
    # Получить максимальное количество заполненных ячеек в первом столбце
    max_row = sheet.max_row

    # Пройти по каждой строке и добавить значения в словарь
    for row in range(4, max_row+1):
            if sheet[f'B{row}'].value != None:
                cell_b = sheet[f'B{row}'].value 
                cell_c = sheet[f'C{row}'].value
                cell_d = sheet[f'D{row}'].value
                cell_e = sheet[f'E{row}'].value
                cell_f = sheet[f'F{row}'].value
                cell_g = sheet[f'G{row}'].value
                if cell_g == '=FALSE()':
                    cell_g= 'ЛОЖЬ'
                if cell_g == '=TRUE()':
                     cell_g= 'ИСТИНА'
            # Добавить значения в словарь
            dictionary[cell_b] = [cell_c, cell_d, cell_e, cell_f, cell_g]
    subjectID= sheet['E2'].value

    url = f"https://suppliers-api.wildberries.ru/content/v2/object/charcs/{subjectID}"

    payload = ""
    headers = {
    'Authorization': 'Bearer eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjMxMDI1djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTcxODMxMDQ5OCwiaWQiOiI0NjcwMGY2MS04ODdjLTQwMTYtYTExNy04MDM3MTE4Nzk3YWQiLCJpaWQiOjY0NDk5NDEsIm9pZCI6OTA3MDUwLCJzIjoyLCJzYW5kYm94IjpmYWxzZSwic2lkIjoiYTA2MTFhYWQtZTQyZC00NmVkLTllZWYtNzYzOTM2YjgzZGYyIiwidWlkIjo2NDQ5OTQxfQ.sG6cJHF4h8NDnNuHo8YmZtTfshmRUFNvBSis3gGYBVNZ9DggrDegHMOnQv4xm-dOeKQyyPS41zPhqV8ZPs11Qg'
    }

    response = requests.request("GET", url, headers=headers, data=payload)

    # print(response.text)
    data=response.json()
    templates = data
    
    js1={}
    if   templates['data'] != None: 
        if isinstance(templates['data'], list):
                    s= templates['data']
                    for t in range(0,len(s)):            
                        s1= s[t]
                        s2=s1['charcID']
                        s3=s1['name']
                        s4=s1['charcType']
                        s5=s1['unitName']
                        if s5 == "":
                              s5= None                        
                        s6=s1['maxCount']
                        s7=s1['required'] 
                        if s7 == False:
                              s7='ЛОЖЬ'
                        if s7 == True:
                              s7='ИСТИНА'
                        js1[s2]=[s3, s4, s5, s6, s7]
        else: 
            print(f'{i} : Не корректный Json')
    else:
          print(f'{i} : Не корректный Json')
   
    def find_differences(dictionary, js1):
    # Нахождение ключей, которые присутствуют только в одном из словарей
        keys_only_in_dict1=[]
        keys_only_in_dict2=[]
        keys_only_in_dict1 = list(set(dictionary.keys()) - set(js1.keys()))
        keys_only_in_dict2 = list(set(js1.keys()) - set(dictionary.keys()))

    # Нахождение различий по значениям для общих ключей
        value_differences = {}
        common_keys = set(dictionary.keys()).intersection(js1.keys())
        for key in common_keys:
            value1 = dictionary[key]
            value2 = js1[key]
            if value1 != value2:
                value_differences[key] = (value1, value2)

        return keys_only_in_dict1, keys_only_in_dict2, value_differences

    # Получение различий между словарями
    keys_only_in_dict1, keys_only_in_dict2, value_differences = find_differences(dictionary, js1)

    # Вывод результатов
    if keys_only_in_dict2 != [] or keys_only_in_dict1 != [] or value_differences !={}:
        print(f'### {i}')
        
        if keys_only_in_dict2 != []:
            notresult.append('NOK')
            print('Отсутствуют:\n')            
            for u in keys_only_in_dict2:
                print(u)
        
        if keys_only_in_dict1 != []:
            notresult.append('NOK')
            print('\nЛишние:\n')
            print(keys_only_in_dict1)
            for u1 in keys_only_in_dict1:        
                print(u1)
        
        if value_differences !={}:
            notresult.append('NOK')
            print("Различия по name, charcType, unitName, maxCount, required:")
            for key, values in value_differences.items():
                value1, value2 = values
                print(f"id атрибута: {key}, Значение файле: {value1}, Значение в WB: {value2}")
        else: 
            print('Различий по параметрам нет \n')

if notresult == []:
      print('Изменений в API WB НЕТ') 
print(len(notresult))     

