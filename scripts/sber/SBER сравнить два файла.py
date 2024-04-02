from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook
from openpyxl.utils import units


wb = load_workbook(f'/home/akozlov/Загрузки/SberTemplates/SBER/Процессоры/Гарнитура Bluetooth для смартфона.xlsx')
wb2 = load_workbook(f'/home/akozlov/Загрузки/SberTemplates/SBER/Процессоры/Гарнитура Bluetooth для смартфона (1).xlsx')
sheet = wb['Гарнитура Bluetooth для смартфо']
sheet2 = wb2['Гарнитура Bluetooth для смартфо']
# mylist1=[]
dates = [x.value for x in sheet.iter_rows(min_col=1,min_row=3,max_row=3).send(None)]
# for hed in dates:
#     mylist1.append(hed) 
    
# mylist2=[]
dates2 = [x.value for x in sheet2.iter_rows(min_col=1,min_row=3,max_row=3).send(None)]
# for hed in dates2:
#     mylist2.append(hed) 
# print(i)

result = list(set(dates) - set(dates2))
result1 = list(set(dates2) - set(dates))
# if 'Артикул производителя (Part Number)*' in result:
#     print(i)
print('Отличие Старого от Нового:')
for i in result:    
    print(f'- {i}')

print('Отличие Нового от Старого:')
for p in result1:    
    print(f'- {p}')
# print(f'Отличие Старого от Нового: {result}')
# print(f'Отличие Нового от Старого: {result1}')